"""
scripts/convert_knowledge.py
-----------------------------
Converts Excel files in knowledge/ to Markdown for RAG indexing.

Each sheet becomes a ## section. Only field-definition rows are kept
(rows where the "Campo" column contains a snake_case identifier).
Actual code-value rows (lookup tables with hundreds of entries) are skipped.

Usage:
    python scripts/convert_knowledge.py
    python scripts/convert_knowledge.py --force   # re-convert even if .md is newer
"""

import argparse
import os
import re

import openpyxl

KNOWLEDGE_DIR = os.path.join(os.path.dirname(__file__), "..", "knowledge")
MAPEO_MAX_LEN = 250


def _clean(val) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val)).strip()


def _is_field_name(val) -> bool:
    """True when val looks like a technical column name (snake_case identifier)."""
    return bool(val and isinstance(val, str) and re.match(r"^[a-z][a-z0-9_]{1,}$", val.strip()))


def _find_campo_position(ws) -> tuple[int, int] | tuple[None, None]:
    """Return (row_idx, col_idx) of the first cell containing 'Campo'."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        for col_idx, val in enumerate(row):
            if val == "Campo":
                return row_idx, col_idx
    return None, None


def _extract_source_systems(row: tuple, campo_col: int) -> dict[str, tuple[int, int]]:
    """
    Given the source-systems row (the one before the header), return
    {system_name: (tipo_col, mapeo_col)} for columns after campo_col + 3.

    Source systems are declared in pairs (TIPO col, MAPEO col). The system
    name cell is at the TIPO col; the MAPEO col is the next one.
    Duplicate system names get a numeric suffix.
    """
    data_start = campo_col + 4
    systems: dict[str, tuple[int, int]] = {}
    seen: dict[str, int] = {}

    i = data_start
    while i < len(row):
        val = row[i]
        if val and isinstance(val, str) and val.strip():
            name = val.strip()
            count = seen.get(name, 0)
            seen[name] = count + 1
            key = name if count == 0 else f"{name} ({count + 1})"
            systems[key] = (i, i + 1)
            i += 2
        else:
            i += 1

    return systems


_HEADER_KEYWORDS = {"TIPO", "MAPEO", "PK", "ENTIDAD/MAESTRO", "CAMPO", "NOMBRE"}


def _extract_inline_systems(header_row: tuple, campo_col: int) -> dict[str, int]:
    """
    For sheets where system names are directly in the header row (no TIPO/MAPEO split).
    Returns {system_name: col_idx}. Ignores structural keywords (TIPO, MAPEO, etc.).
    """
    data_start = campo_col + 4
    systems: dict[str, int] = {}
    seen: dict[str, int] = {}
    for i in range(data_start, len(header_row)):
        val = header_row[i]
        if (
            val
            and isinstance(val, str)
            and val.strip().upper() == val.strip()
            and val.strip() not in _HEADER_KEYWORDS
        ):
            name = val.strip()
            count = seen.get(name, 0)
            seen[name] = count + 1
            key = name if count == 0 else f"{name} ({count + 1})"
            systems[key] = i
    return systems


def _find_systems_row(
    all_rows: list, header_row_idx: int, campo_col: int
) -> tuple | None:
    """
    Search all rows before the header for one that has source-system names
    (uppercase words, not structural keywords) at cols >= campo_col+4.
    Returns the row tuple or None.
    """
    _SYSTEM_KEYWORDS = {"TIPO", "MAPEO", "PK", "ENTIDAD/MAESTRO", "CAMPO", "NOMBRE", "RESUMEN"}
    for row in all_rows[:header_row_idx]:
        candidates = [
            v
            for i, v in enumerate(row)
            if i >= campo_col + 4
            and v
            and isinstance(v, str)
            and v.strip().upper() == v.strip()
            and v.strip() not in _SYSTEM_KEYWORDS
            and len(v.strip()) > 1
        ]
        if candidates:
            return row
    return None


def _resumen_sheet_to_md(ws) -> str:
    lines = ["## Resumen de entidades\n"]
    lines.append("| Entidad | Tipo de entidad | Descripción | Dataset |")
    lines.append("|---------|----------------|-------------|---------|")
    for row in ws.iter_rows(values_only=True):
        if not row[0] or not isinstance(row[0], str) or row[0] == "Entidad":
            continue
        cols = [_clean(row[i]) if i < len(row) else "" for i in range(4)]
        lines.append(f"| {cols[0]} | {cols[1]} | {cols[2]} | {cols[3]} |")
    return "\n".join(lines) + "\n"


def _dataset_sheet_to_md(sheet_name: str, ws) -> str:
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return f"## {sheet_name}\n\n(Hoja vacía)\n"

    header_row_idx, campo_col = _find_campo_position(ws)
    if header_row_idx is None:
        return f"## {sheet_name}\n\n(Sin estructura de campos reconocida)\n"

    header_row = all_rows[header_row_idx]

    # Description: look for a non-empty string cell before the header row
    description = ""
    for row in all_rows[1:header_row_idx]:
        for val in row:
            if val and isinstance(val, str) and len(val.strip()) > 20:
                description = _clean(val)
                break
        if description:
            break

    # Determine if systems are in a separate pre-header row or inline in the header
    systems_row = _find_systems_row(all_rows, header_row_idx, campo_col)

    use_inline = False
    if systems_row is not None:
        systems = _extract_source_systems(systems_row, campo_col)
    else:
        inline = _extract_inline_systems(header_row, campo_col)
        if inline:
            use_inline = True
            systems = {name: (col, None) for name, col in inline.items()}
        else:
            systems = {}

    # Build markdown
    lines = [f"## {sheet_name}\n"]
    if description:
        lines.append(f"{description}\n")

    # Table header
    sys_names = list(systems.keys())
    if sys_names:
        header_cols = "| Campo | Nombre | Descripción funcional | Maestro | " + " | ".join(sys_names) + " |"
        sep_cols = "|---|---|---|---|" + "---|" * len(sys_names)
    else:
        header_cols = "| Campo | Nombre | Descripción funcional | Maestro |"
        sep_cols = "|---|---|---|---|"
    lines.append(header_cols)
    lines.append(sep_cols)

    field_count = 0
    for row in all_rows[header_row_idx + 1:]:
        if not row or len(row) <= campo_col:
            continue
        campo_val = row[campo_col]
        if not _is_field_name(campo_val):
            continue

        campo = _clean(campo_val)
        nombre = _clean(row[campo_col + 1]) if len(row) > campo_col + 1 else ""
        desc = _clean(row[campo_col + 2]) if len(row) > campo_col + 2 else ""
        maestro = _clean(row[campo_col + 3]) if len(row) > campo_col + 3 else ""

        sys_cells = []
        for sys_name, cols in systems.items():
            tipo_col, mapeo_col = cols
            if use_inline:
                val = _clean(row[tipo_col]) if tipo_col < len(row) else ""
                sys_cells.append(val[:MAPEO_MAX_LEN] if val else "—")
            else:
                tipo = _clean(row[tipo_col]) if tipo_col is not None and tipo_col < len(row) else ""
                mapeo = _clean(row[mapeo_col]) if mapeo_col is not None and mapeo_col < len(row) else ""
                if tipo and mapeo:
                    cell = f"**{tipo}**: {mapeo}"
                elif tipo:
                    cell = tipo
                else:
                    cell = mapeo or "—"
                sys_cells.append(cell[:MAPEO_MAX_LEN])

        sys_str = " | ".join(sys_cells)
        row_str = f"| {campo} | {nombre} | {desc} | {maestro} |"
        if sys_cells:
            row_str += f" {sys_str} |"
        lines.append(row_str)
        field_count += 1

    if field_count == 0:
        lines.append("*(Sin definiciones de campos encontradas)*")

    lines.append(f"\n*{field_count} campos definidos*\n")
    return "\n".join(lines)


def convert_xlsx(xlsx_path: str, output_dir: str) -> str:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    output_path = os.path.join(output_dir, f"{base.lower().replace(' ', '_')}.md")

    title = base.replace("_", " ").title()
    sections = [
        f"# {title}\n\n",
        f"*Generado desde `{os.path.basename(xlsx_path)}`*\n\n",
    ]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.lower() == "resumen":
            sections.append(_resumen_sheet_to_md(ws))
        else:
            sections.append(_dataset_sheet_to_md(sheet_name, ws))

    content = "\n".join(sections)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    return output_path


def main():
    parser = argparse.ArgumentParser(description="Convert Excel knowledge files to Markdown")
    parser.add_argument(
        "--force", action="store_true",
        help="Re-convert even if the .md file is newer than the .xlsx"
    )
    args = parser.parse_args()

    knowledge_dir = os.path.abspath(KNOWLEDGE_DIR)
    if not os.path.isdir(knowledge_dir):
        print(f"[!] Directory not found: {knowledge_dir}")
        return

    converted = 0
    skipped = 0
    for fname in sorted(os.listdir(knowledge_dir)):
        if not fname.endswith(".xlsx"):
            continue
        xlsx_path = os.path.join(knowledge_dir, fname)
        base = os.path.splitext(fname)[0].lower().replace(" ", "_")
        md_path = os.path.join(knowledge_dir, f"{base}.md")

        if not args.force and os.path.exists(md_path):
            if os.path.getmtime(md_path) >= os.path.getmtime(xlsx_path):
                print(f"  skip  {fname}  (already up to date)")
                skipped += 1
                continue

        out = convert_xlsx(xlsx_path, knowledge_dir)
        size_kb = os.path.getsize(out) // 1024
        print(f"  done  {fname} → {os.path.basename(out)} ({size_kb} KB)")
        converted += 1

    print(f"\n{converted} converted, {skipped} skipped.")


if __name__ == "__main__":
    main()
